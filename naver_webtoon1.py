from bs4 import BeautifulSoup as bs
from pprint import pprint
import requests, re, os
from urllib.request import  urlretrieve
from openpyxl.drawing.image import Image
from openpyxl import Workbook

wb = Workbook()
sheet1 = wb.active
sheet1.title = '네이버 웹툰 완결편'

try:
    if not (os.path.isdir('image')):
        os.mkdir(os.path.join('image'))
except OSError as e:
    if e.errno != errno.EEXITS:
        print("폴더생성실패")
        exit()
url = 'https://comic.naver.com/webtoon/finish.nhn'
html = requests.get(url)
soup = bs(html.text, 'html.parser')
all_data = soup.find('div', {'class':'list_area'})

data_list = all_data.findAll('li')

col = 1
row = 1
for data in data_list:
    img_list = data.find('img')
    img_src = img_list['src']
    a_list = data.find('a')
    title = a_list['title']
    title = re.sub('[^0-9a-zA-Zㄱ-힗]', '', title)
    link = "https//commic.naver.com" + a_list['href']
    strong = data.find('strong').text

#   urlretrieve(img_src, './image/'+title+'.gif')
    img_title = './image/'+title+'.gif'
    img_file = Image(img_title)
    print(img_title)
    img_file.anchor= 'C10'
    #cell = sheet1.cell(row=10, column=1)
    sheet1.add_image(img_file)
#    sheet1.cell()
#    col = col + 1
    row = row + 1
    break
wb.save("./webtoon.xlsx")
    #print(title, strong, link)
'''
data2 = data1.findAll('dd')
pprint(data2)
data3 = data2[0].find('span')
print(data3.text)
'''